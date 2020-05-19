from pymongo import MongoClient 
from openpyxl import load_workbook
from decouple import config

try: 
	client = MongoClient(config('MONGO_CONNECTION_STRING_FOR_TESTING')) 
	envDb = client.environment
	collection = envDb.cities
	print("Connected successfully!") 
except: 
	print("Could not connect to MongoDB") 
	

workbook = load_workbook(filename="Jones-Kammen-2014-Zip-City-County-Results.xlsx")
workbook.active = 2
sheet = workbook.active
idNum = 1
# skip 1140 for {city: "CA"}, line 26672 is the last CA

# if number_of_cities is 0, the script will run for all cities
# if number_of_cities is given a number other than 0, the script will run for the number of cities designated

number_of_cities = 0

if(number_of_cities == 0):
	number_of_cities = 26784

number_of_cities+=1

#26785 total rows

for row in sheet.iter_rows(min_row=2, max_row=number_of_cities, min_col=1, values_only=True):
	state = row[0]
	county = row[1]
	city = row[2]
	co2 = row[16]
	co2_household = row[14]
	transport = row[9]
	housing = row[10]
	food = row [11]
	goods = row[12]
	services = row[13]
	electricity = row[5]
	natGas = row[6]
	oil = row[7]
	milesTraveled = row[8]

	location = {
		"_id":idNum, 
        "state":state, 
        "county":county,
		"name":city,
		"CO2":co2,
		"CO2_Emissions_Per_Household": co2_household,
		"Transport": transport,
		"Housing": housing,
		"Food": food,
		"Goods": goods,
		"Services": services,
		"Electricity": electricity,
		"Natural_Gas": natGas,
		"Fuel_Oil": oil,
		"Vehicle_Miles_Traveled": milesTraveled
    }
	collection.insert_one(location)
	idNum+=1

print("Complete")

	









# https://www.joe0.com/2019/04/20/python-tutorial-reading-writing-excel-files-data-gathering-by-web-scraping-google/
# https://www.geeksforgeeks.org/mongodb-python-insert-update-data/
# https://api.mongodb.com/python/current/tutorial.html